"""
Excel report writer
"""

import logging
from pathlib import Path
from datetime import datetime
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ...core.models.report import PaymentReport
from ...core.models.payment import PaymentStatus
from ...infrastructure.localization.i18n import LocalizationManager
from ...core.exceptions import FileProcessingError


class ExcelReportWriter:
    """
    Writer for generating Excel payment reports
    """
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
    
    def write_report(self, 
                    report: PaymentReport, 
                    output_path: Path,
                    i18n: LocalizationManager):
        """
        Write payment report to Excel file
        
        Args:
            report: Payment report data
            output_path: Output file path
            i18n: Localization manager for translations
            
        Raises:
            FileProcessingError: If writing fails
        """
        try:
            workbook = Workbook()
            
            # Create main report worksheet
            self._create_payments_worksheet(workbook, report, i18n)
            
            # Create summary worksheet
            self._create_summary_worksheet(workbook, report, i18n)
            
            # Save workbook
            workbook.save(output_path)
            self.logger.info(f"Excel report saved to {output_path}")
            
        except Exception as e:
            raise FileProcessingError(f"Failed to write Excel report: {e}", str(output_path))
    
    def _create_payments_worksheet(self, 
                                 workbook: Workbook, 
                                 report: PaymentReport,
                                 i18n: LocalizationManager):
        """Create main payments worksheet"""
        
        # Use default worksheet
        ws = workbook.active
        ws.title = i18n.get("report.worksheet.payments", default="Payments")
        
        # Add report metadata at the top
        current_row = 1
        
        # Report generation date
        generation_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.cell(row=current_row, column=1, value="Report date:")
        ws.cell(row=current_row, column=2, value=generation_date)
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        current_row += 1
        
        # Analysis date
        analysis_date_str = report.analysis_date.strftime("%Y-%m-%d") if report.analysis_date else "Not specified"
        ws.cell(row=current_row, column=1, value="Analysis date:")
        ws.cell(row=current_row, column=2, value=analysis_date_str)
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        current_row += 1
        
        # Empty row for separation
        current_row += 1
        
        # Define headers
        headers = [
            i18n.get("report.header.garage", default="Garage"),
            i18n.get("report.header.amount", default="Amount (RUB)"),
            i18n.get("report.header.expected_date", default="Expected Date"),
            i18n.get("report.header.actual_date", default="Actual Date"),
            i18n.get("report.header.status", default="Status"),
            i18n.get("report.header.days_overdue", default="Days Overdue"),
            i18n.get("report.header.notes", default="Notes")
        ]
        
        # Write headers
        header_row = current_row
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Write payment data (starting from row after headers)
        data_start_row = header_row + 1
        for i, payment in enumerate(report.payments):
            row = data_start_row + i
            # Garage ID
            ws.cell(row=row, column=1, value=payment.garage_id)
            
            # Amount
            ws.cell(row=row, column=2, value=float(payment.amount))
            
            # Expected date
            ws.cell(row=row, column=3, value=payment.expected_date.strftime("%Y-%m-%d"))
            
            # Actual date
            actual_date_str = payment.actual_date.strftime("%Y-%m-%d") if payment.actual_date else ""
            ws.cell(row=row, column=4, value=actual_date_str)
            
            # Status
            status_text = self._get_status_text(payment.status, i18n)
            status_cell = ws.cell(row=row, column=5, value=status_text)
            status_cell.fill = self._get_status_color(payment.status)
            
            # Days overdue - show for received, overdue, and pending payments
            if payment.status in [PaymentStatus.RECEIVED, PaymentStatus.OVERDUE, PaymentStatus.PENDING]:
                overdue_value = payment.days_overdue if payment.days_overdue is not None else 0
            else:
                overdue_value = ""
            ws.cell(row=row, column=6, value=overdue_value)
            
            # Notes
            ws.cell(row=row, column=7, value=payment.notes)
        
        # Auto-size columns
        self._auto_size_columns(ws)
        
        # Add borders for data table only (excluding metadata at top)
        total_data_rows = len(report.payments) + 1  # +1 for header
        self._add_borders(ws, total_data_rows, len(headers), header_row)
    
    def _create_summary_worksheet(self, 
                                workbook: Workbook, 
                                report: PaymentReport,
                                i18n: LocalizationManager):
        """Create summary worksheet"""
        
        ws = workbook.create_sheet(title=i18n.get("report.worksheet.summary", default="Summary"))
        
        summary = report.summary
        
        # Report header
        ws.cell(row=1, column=1, value=i18n.get("summary.title", default="Payment Summary")).font = Font(size=16, bold=True)
        ws.cell(row=2, column=1, value=f"Generated: {report.generated_at.strftime('%Y-%m-%d %H:%M')}")
        ws.cell(row=3, column=1, value=f"Analysis Date: {report.analysis_date.strftime('%Y-%m-%d')}")
        
        # Statistics
        row = 5
        stats = [
            (i18n.get("summary.total_garages", default="Total Garages"), summary.total_garages),
            (i18n.get("summary.received", default="Received"), summary.received_count),
            (i18n.get("summary.overdue", default="Overdue"), summary.overdue_count),
            (i18n.get("summary.pending", default="Pending"), summary.pending_count),
            (i18n.get("summary.not_due", default="Not Due"), summary.not_due_count),
            (i18n.get("summary.unclear", default="Unclear"), summary.unclear_count),
            ("", ""),  # Empty row
            (i18n.get("summary.collection_rate", default="Collection Rate"), f"{summary.collection_rate:.1f}%"),
            (i18n.get("summary.expected_amount", default="Expected Amount"), f"{summary.total_expected:.2f} RUB"),
            (i18n.get("summary.received_amount", default="Received Amount"), f"{summary.total_received:.2f} RUB")
        ]
        
        for label, value in stats:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        # Notes section
        if report.notes:
            row += 1
            ws.cell(row=row, column=1, value=i18n.get("summary.notes_header", default="Notes")).font = Font(bold=True)
            row += 1
            
            for note in report.notes:
                ws.cell(row=row, column=1, value=f"• {note}")
                row += 1
        
        # Auto-size columns
        self._auto_size_columns(ws)
    
    def _get_status_text(self, status: PaymentStatus, i18n: LocalizationManager) -> str:
        """Get localized status text"""
        status_map = {
            PaymentStatus.RECEIVED: i18n.get("status.received", default="Received"),
            PaymentStatus.OVERDUE: i18n.get("status.overdue", default="Overdue"),
            PaymentStatus.PENDING: i18n.get("status.pending", default="Pending"),
            PaymentStatus.NOT_DUE: i18n.get("status.not_due", default="Not Due"),
            PaymentStatus.UNCLEAR: i18n.get("status.unclear", default="Unclear")
        }
        return status_map.get(status, status.value)
    
    def _get_status_color(self, status: PaymentStatus) -> PatternFill:
        """Get color for status cell"""
        color_map = {
            PaymentStatus.RECEIVED: PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),  # Light green
            PaymentStatus.OVERDUE: PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid"),   # Light red
            PaymentStatus.PENDING: PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid"),   # Light yellow
            PaymentStatus.NOT_DUE: PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid"),   # Light purple
            PaymentStatus.UNCLEAR: PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")    # Orange
        }
        return color_map.get(status, PatternFill())
    
    def _auto_size_columns(self, worksheet):
        """Auto-size worksheet columns"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    cell_length = len(str(cell.value or ''))
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
            
            # Set column width with some padding
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _add_borders(self, worksheet, num_rows: int, num_cols: int, start_row: int = 1):
        """Add borders to table"""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(start_row, start_row + num_rows):
            for col in range(1, num_cols + 1):
                worksheet.cell(row=row, column=col).border = thin_border
