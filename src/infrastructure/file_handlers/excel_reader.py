"""
Excel file reading utilities
"""

import logging
from pathlib import Path
from typing import List, Dict, Any, Optional, Iterator

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet

from ...core.exceptions import FileProcessingError


class ExcelReader:
    """
    Utility class for reading Excel files
    """
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
    
    def read_workbook(self, file_path: Path, read_only: bool = True, data_only: bool = True):
        """
        Read Excel workbook
        
        Args:
            file_path: Path to Excel file
            read_only: Open in read-only mode
            data_only: Read cell values instead of formulas
            
        Returns:
            Openpyxl workbook object
            
        Raises:
            FileProcessingError: If file cannot be read
        """
        try:
            return load_workbook(file_path, read_only=read_only, data_only=data_only)
        except InvalidFileException as e:
            raise FileProcessingError(f"Invalid Excel file format: {e}", str(file_path))
        except Exception as e:
            raise FileProcessingError(f"Failed to read Excel file: {e}", str(file_path))
    
    def read_workbook_safe(self, file_path: Path, read_only: bool = True, data_only: bool = True):
        """
        Read Excel workbook with guaranteed cleanup
        
        Args:
            file_path: Path to Excel file
            read_only: Open in read-only mode
            data_only: Read cell values instead of formulas
            
        Returns:
            Openpyxl workbook object
            
        Raises:
            FileProcessingError: If file cannot be read
        """
        workbook = None
        try:
            workbook = load_workbook(file_path, read_only=read_only, data_only=data_only)
            return workbook
        except InvalidFileException as e:
            if workbook:
                try:
                    workbook.close()
                except:
                    pass
            raise FileProcessingError(f"Invalid Excel file format: {e}", str(file_path))
        except Exception as e:
            if workbook:
                try:
                    workbook.close()
                except:
                    pass
            raise FileProcessingError(f"Failed to read Excel file: {e}", str(file_path))
    
    def get_worksheet_data(self, 
                          worksheet: Worksheet, 
                          min_row: int = 1, 
                          max_row: Optional[int] = None,
                          min_col: int = 1,
                          max_col: Optional[int] = None) -> Iterator[List[Any]]:
        """
        Get data from worksheet as iterator of rows
        
        Args:
            worksheet: Worksheet to read from
            min_row: Minimum row number
            max_row: Maximum row number (None for all)
            min_col: Minimum column number
            max_col: Maximum column number (None for all)
            
        Yields:
            List of cell values for each row
        """
        for row in worksheet.iter_rows(
            min_row=min_row, 
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            values_only=True
        ):
            yield list(row)
    
    def find_header_row(self, 
                       worksheet: Worksheet, 
                       header_keywords: List[str], 
                       max_search_rows: int = 20) -> Optional[int]:
        """
        Find row containing header keywords
        
        Args:
            worksheet: Worksheet to search in
            header_keywords: List of keywords to look for
            max_search_rows: Maximum number of rows to search
            
        Returns:
            Row number (1-based) or None if not found
        """
        for row_num, row in enumerate(worksheet.iter_rows(min_row=1, max_row=max_search_rows), 1):
            row_text = ' '.join(str(cell.value or '').lower() for cell in row)
            
            # Check if enough keywords are present
            found_keywords = sum(1 for keyword in header_keywords if keyword.lower() in row_text)
            if found_keywords >= len(header_keywords) // 2:  # At least half the keywords
                return row_num
        
        return None
    
    def extract_cell_value(self, cell_value: Any, expected_type: type = str) -> Any:
        """
        Extract and convert cell value to expected type
        
        Args:
            cell_value: Raw cell value
            expected_type: Expected Python type
            
        Returns:
            Converted value or None if conversion fails
        """
        if cell_value is None:
            return None
        
        try:
            if expected_type == str:
                return str(cell_value).strip()
            elif expected_type == int:
                return int(float(cell_value))  # Handle string numbers
            elif expected_type == float:
                return float(cell_value)
            else:
                return cell_value
        except (ValueError, TypeError):
            self.logger.warning(f"Could not convert cell value '{cell_value}' to {expected_type.__name__}")
            return None
    
    def is_empty_row(self, row: List[Any]) -> bool:
        """
        Check if row is empty (all None values)
        
        Args:
            row: List of cell values
            
        Returns:
            True if row is empty
        """
        return all(value is None or str(value).strip() == '' for value in row)
    
    def validate_file_exists(self, file_path: Path) -> bool:
        """
        Validate that file exists and is readable
        
        Args:
            file_path: Path to file
            
        Returns:
            True if file is valid
        """
        return file_path.exists() and file_path.is_file() and file_path.suffix.lower() in ['.xlsx', '.xls']
    
    def read_file_data_safe(self, file_path: Path) -> List[List[Any]]:
        """
        Read Excel file data with guaranteed cleanup
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            List of rows with cell values
            
        Raises:
            FileProcessingError: If file cannot be read
        """
        workbook = None
        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            worksheet = workbook.active
            
            data = []
            for row in worksheet.iter_rows(values_only=True):
                data.append(list(row))
            
            return data
            
        except InvalidFileException as e:
            raise FileProcessingError(f"Invalid Excel file format: {e}", str(file_path))
        except Exception as e:
            raise FileProcessingError(f"Failed to read Excel file: {e}", str(file_path))
        finally:
            if workbook:
                try:
                    workbook.close()
                except Exception as e:
                    self.logger.warning(f"Error closing workbook: {e}")
