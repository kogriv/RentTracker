"""
Sberbank statement parser for Excel format
"""

import logging
import re
from pathlib import Path
from typing import List, Optional
from decimal import Decimal
from datetime import datetime, date

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from ..base.parser_interface import StatementParser
from ...core.models.transaction import Transaction
from ...core.models.payment_period import PaymentPeriod
from ...core.exceptions import ParseError


class SberbankStatementParser(StatementParser):
    """
    Parser for Sberbank bank statements in Excel format
    
    Expected format based on the specification:
    - Column A: Date and time (дд.мм.гггг чч:мм)
    - Column B: Time (may be separate)
    - Column C: Usually empty or additional info
    - Column D: Operation category (Перевод на карту, Перевод СБП, etc.)
    - Column E: Amount (+X XXX,XX for incoming)
    """
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        
        # Regex patterns for parsing
        self.date_pattern = re.compile(r'(\d{2}\.\d{2}\.\d{4})')
        self.amount_pattern = re.compile(r'\+.*?(\d[\d\s,]*\d|\d),\d{2}')
        self.incoming_keywords = ['перевод', 'сбп', 'карту']
        
        # Pattern for period detection
        self.period_pattern = re.compile(r'итого\s+по\s+операциям\s+с\s+(\d{2}\.\d{2}\.\d{4})\s+по\s+(\d{2}\.\d{2}\.\d{4})', re.IGNORECASE)
    
    def parse_transactions(self, source: Path) -> List[Transaction]:
        """
        Parse transactions from Sberbank Excel statement
        
        Args:
            source: Path to Excel file
            
        Returns:
            List of parsed transactions
            
        Raises:
            ParseError: If parsing fails
        """
        if not self.validate_source(source):
            raise ParseError(f"Invalid Excel file: {source}")
        
        workbook = None
        try:
            workbook = load_workbook(source, read_only=True, data_only=True)
            worksheet = workbook.active
            
            transactions = []
            
            for row_num, row in enumerate(worksheet.iter_rows(min_row=1), 1):
                try:
                    # Skip empty rows
                    if all(cell.value is None for cell in row):
                        continue
                    
                    # Parse transaction from row
                    transaction = self._parse_transaction_row(row, row_num)
                    if transaction:
                        transactions.append(transaction)
                
                except Exception as e:
                    self.logger.debug(f"Error parsing row {row_num}: {e}")
                    continue
            
            # Filter for incoming transactions only
            incoming_transactions = [t for t in transactions if t.amount > 0]
            
            self.logger.info(f"Parsed {len(incoming_transactions)} incoming transactions from {len(transactions)} total")
            return incoming_transactions
            
        except InvalidFileException as e:
            raise ParseError(f"Invalid Excel file format: {e}", str(source))
        except Exception as e:
            raise ParseError(f"Failed to parse Excel file: {e}", str(source))
        finally:
            if workbook:
                try:
                    workbook.close()
                except Exception as e:
                    self.logger.warning(f"Error closing workbook: {e}")
    
    def validate_source(self, source: Path) -> bool:
        """
        Validate if source file can be parsed
        
        Args:
            source: Path to source file
            
        Returns:
            True if file can be parsed
        """
        if not source.exists():
            return False
        
        if source.suffix.lower() not in ['.xlsx', '.xls']:
            return False
        
        try:
            # Try to open the file and look for Sberbank patterns
            workbook = load_workbook(source, read_only=True, data_only=True)
            worksheet = workbook.active
            
            # Check first few rows for expected patterns
            for row in worksheet.iter_rows(min_row=1, max_row=20):
                if self._looks_like_sberbank_row(row):
                    workbook.close()
                    return True
            
            workbook.close()
            return False
            
        except Exception:
            return False
    
    def get_supported_formats(self) -> List[str]:
        """Get list of supported file formats"""
        return ['.xlsx', '.xls']
    
    def _looks_like_sberbank_row(self, row) -> bool:
        """Check if row looks like a Sberbank transaction row"""
        if len(row) < 2:
            return False
        
        # Check for date pattern in first column
        first_cell = str(row[0].value or '')
        if self.date_pattern.search(first_cell):
            return True
        
        # Check for amount pattern in any column
        for cell in row:
            cell_text = str(cell.value or '')
            if '+' in cell_text and (',' in cell_text or '.' in cell_text):
                return True
        
        return False
    
    def _parse_transaction_row(self, row, row_num: int) -> Optional[Transaction]:
        """
        Parse a single transaction row
        
        Args:
            row: Excel row cells
            row_num: Row number for error reporting
            
        Returns:
            Transaction object or None if invalid
        """
        # Extract data from row
        transaction_date = self._extract_date(row)
        amount = self._extract_amount(row)
        category = self._extract_category(row)
        
        # Must have date and amount
        if not transaction_date or not amount:
            return None
        
        # Only process incoming transactions (positive amounts)
        if amount <= 0:
            return None
        
        # Must be transfer-related transaction
        if not self._is_transfer_category(category):
            return None
        
        return Transaction(
            date=transaction_date,
            amount=amount,
            category=category,
            source=f"sberbank_statement_row_{row_num}",
            description=f"Parsed from row {row_num}"
        )
    
    def _extract_date(self, row) -> Optional[date]:
        """Extract date from row"""
        # Look in first few columns for date
        for i in range(min(3, len(row))):
            cell_value = row[i].value
            if not cell_value:
                continue
            
            # Handle datetime objects
            if isinstance(cell_value, datetime):
                return cell_value.date()
            
            # Handle string dates
            cell_text = str(cell_value)
            date_match = self.date_pattern.search(cell_text)
            if date_match:
                try:
                    date_str = date_match.group(1)
                    return datetime.strptime(date_str, '%d.%m.%Y').date()
                except ValueError:
                    continue
        
        return None
    
    def _extract_amount(self, row) -> Optional[Decimal]:
        """Extract amount from row"""
        # Look in all columns for amount (usually column E, but be flexible)
        for cell in row:
            if not cell.value:
                continue
            
            cell_text = str(cell.value)
            
            # Look for positive amount pattern
            if '+' not in cell_text:
                continue
            
            amount_match = self.amount_pattern.search(cell_text)
            if amount_match:
                try:
                    # Extract numeric part
                    amount_str = amount_match.group(0)
                    # Clean up: remove +, spaces, replace comma with dot
                    amount_str = amount_str.replace('+', '').replace(' ', '').replace(',', '.')
                    return Decimal(amount_str)
                except (ValueError, TypeError):
                    continue
            
            # Fallback: try to parse directly if it looks like a positive number
            if cell_text.startswith('+'):
                try:
                    clean_amount = cell_text.replace('+', '').replace(' ', '').replace(',', '.')
                    # Check if it's a valid number
                    if re.match(r'^\d+\.?\d*$', clean_amount):
                        return Decimal(clean_amount)
                except (ValueError, TypeError):
                    continue
        
        return None
    
    def _extract_category(self, row) -> str:
        """Extract transaction category from row"""
        # Look for category in column D or nearby columns
        for i in range(len(row)):
            cell_value = row[i].value
            if not cell_value:
                continue
            
            cell_text = str(cell_value).lower()
            
            # Check if this looks like a category
            if any(keyword in cell_text for keyword in self.incoming_keywords):
                return str(row[i].value)
        
        return "Unknown Transfer"
    
    def _is_transfer_category(self, category: str) -> bool:
        """Check if category indicates a transfer transaction"""
        if not category:
            return False
        
        category_lower = category.lower()
        return any(keyword in category_lower for keyword in self.incoming_keywords)
    
    def extract_payment_period(self, source: Path) -> Optional[PaymentPeriod]:
        """
        Extract payment period from Sberbank statement
        
        Looks for lines like "Итого по операциям с 01.05.2025 по 12.06.2025"
        
        Args:
            source: Path to Excel file
            
        Returns:
            PaymentPeriod if found, None otherwise
        """
        if not self.validate_source(source):
            return None
        
        workbook = None
        try:
            workbook = load_workbook(source, read_only=True, data_only=True)
            worksheet = workbook.active
            
            for row_num, row in enumerate(worksheet.iter_rows(min_row=1), 1):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        cell_text = cell.value.strip().lower()
                        
                        # Search for period pattern
                        match = self.period_pattern.search(cell_text)
                        if match:
                            start_date_str = match.group(1)
                            end_date_str = match.group(2)
                            
                            try:
                                start_date = datetime.strptime(start_date_str, '%d.%m.%Y').date()
                                end_date = datetime.strptime(end_date_str, '%d.%m.%Y').date()
                                
                                period = PaymentPeriod(
                                    start_date=start_date,
                                    end_date=end_date,
                                    source_text=cell.value.strip()
                                )
                                
                                self.logger.info(f"Found payment period: {period}")
                                return period
                                
                            except ValueError as e:
                                self.logger.warning(f"Failed to parse dates from '{cell.value}': {e}")
                                continue
            
            self.logger.warning(f"No payment period found in {source}")
            return None
            
        except Exception as e:
            self.logger.error(f"Error extracting payment period from {source}: {e}")
            return None
        finally:
            if workbook:
                try:
                    workbook.close()
                except Exception as e:
                    self.logger.warning(f"Error closing workbook: {e}")
