"""
Excel garage registry parser
"""

import logging
from pathlib import Path
from typing import List, Dict, Any
from decimal import Decimal
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from ..base.parser_interface import GarageRegistryParser
from ...core.exceptions import ParseError, ValidationError


class ExcelGarageParser(GarageRegistryParser):
    """
    Parser for garage registry in Excel format
    """
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
    
    def parse_garages(self, source: Path) -> List[Dict[str, Any]]:
        """
        Parse garage registry from Excel file
        
        Expected format:
        - Column A: Garage ID (Гараж)
        - Column B: Monthly rent amount (Сумма)
        - Column C: Start date (Первоначальная дата)
        
        Args:
            source: Path to Excel file
            
        Returns:
            List of garage data dictionaries
            
        Raises:
            ParseError: If parsing fails
        """
        if not self.validate_source(source):
            raise ParseError(f"Invalid Excel file: {source}")
        
        workbook = None
        try:
            workbook = load_workbook(source, read_only=True, data_only=True)
            worksheet = workbook.active
            
            garages = []
            headers_found = False
            
            for row_num, row in enumerate(worksheet.iter_rows(min_row=1), 1):
                try:
                    # Skip empty rows
                    if all(cell.value is None for cell in row):
                        continue
                    
                    # Look for headers first
                    if not headers_found:
                        if self._is_header_row(row):
                            headers_found = True
                            self.logger.debug(f"Found headers at row {row_num}")
                            continue
                        else:
                            # Try to parse as data if no headers found yet
                            pass
                    
                    # Parse data row
                    garage_data = self._parse_garage_row(row, row_num)
                    if garage_data:
                        garages.append(garage_data)
                
                except Exception as e:
                    self.logger.warning(f"Error parsing row {row_num}: {e}")
                    continue
            
            if not garages:
                raise ParseError("No valid garage data found in file", str(source))
            
            self.logger.info(f"Successfully parsed {len(garages)} garages from {source}")
            return garages
            
        except InvalidFileException as e:
            raise ParseError(f"Invalid Excel file format: {e}", str(source))
        except Exception as e:
            raise ParseError(f"Failed to parse Excel file: {e}", str(source))
        finally:
            if workbook:
                try:
                    workbook.close()
                except Exception as e:
                    self.logger.warning(f"Error closing workbook: {e}")
    
    def validate_source(self, source: Path) -> bool:
        """
        Validate if source file can be parsed
        
        Args:
            source: Path to source file
            
        Returns:
            True if file can be parsed
        """
        if not source.exists():
            return False
        
        if source.suffix.lower() not in ['.xlsx', '.xls']:
            return False
        
        try:
            # Try to open the file
            workbook = load_workbook(source, read_only=True)
            workbook.close()
            return True
        except Exception:
            return False
    
    def _is_header_row(self, row) -> bool:
        """Check if row contains header information"""
        if len(row) < 3:
            return False
        
        # Look for expected header keywords
        header_keywords = ['гараж', 'сумма', 'дата', 'первоначальная']
        row_text = ' '.join(str(cell.value or '').lower() for cell in row[:3])
        
        return any(keyword in row_text for keyword in header_keywords)
    
    def _parse_garage_row(self, row, row_num: int) -> Dict[str, Any]:
        """
        Parse a single garage data row
        
        Args:
            row: Excel row cells
            row_num: Row number for error reporting
            
        Returns:
            Garage data dictionary or None if invalid
        """
        if len(row) < 3:
            return None
        
        # Extract cell values
        garage_id_cell = row[0].value
        amount_cell = row[1].value
        date_cell = row[2].value
        
        # Skip if essential data is missing
        if garage_id_cell is None or amount_cell is None or date_cell is None:
            return None
        
        try:
            # Parse garage ID
            garage_id = str(garage_id_cell).strip()
            if not garage_id:
                return None
            
            # Parse amount
            if isinstance(amount_cell, (int, float)):
                amount = Decimal(str(amount_cell))
            else:
                # Try to parse string amount
                amount_str = str(amount_cell).replace(',', '.').strip()
                amount = Decimal(amount_str)
            
            if amount <= 0:
                self.logger.warning(f"Invalid amount {amount} at row {row_num}")
                return None
            
            # Parse date
            if isinstance(date_cell, datetime):
                start_date = date_cell.date()
            else:
                # Try to parse string date
                date_str = str(date_cell).strip()
                # Common date formats
                for fmt in ['%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d']:
                    try:
                        start_date = datetime.strptime(date_str, fmt).date()
                        break
                    except ValueError:
                        continue
                else:
                    self.logger.warning(f"Could not parse date '{date_str}' at row {row_num}")
                    return None
            
            return {
                'id': garage_id,
                'monthly_rent': amount,
                'start_date': start_date,
                'payment_day': start_date.day
            }
            
        except (ValueError, TypeError) as e:
            self.logger.warning(f"Error parsing garage data at row {row_num}: {e}")
            return None
    
    def get_supported_formats(self) -> List[str]:
        """Get list of supported file formats"""
        return ['.xlsx', '.xls']
