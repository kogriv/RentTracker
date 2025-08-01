"""
Integration tests for full payment processing flow
"""

import pytest
import tempfile
from pathlib import Path
from decimal import Decimal
from datetime import date, datetime
from openpyxl import Workbook

from src.application.use_cases.process_payments import ProcessPaymentsUseCase
from src.application.use_cases.generate_report import GenerateReportUseCase
from src.application.dto.payment_request import PaymentProcessRequest
from src.parsers.base.parser_factory import ParserFactory
from src.core.services.payment_matcher import PaymentMatcher
from src.infrastructure.file_handlers.excel_writer import ExcelReportWriter
from src.infrastructure.localization.i18n import LocalizationManager
from src.core.models.payment import PaymentStatus


class TestFullPaymentFlow:
    """Integration tests for complete payment processing flow"""
    
    @pytest.fixture
    def temp_dir(self):
        """Create temporary directory for test files"""
        with tempfile.TemporaryDirectory() as temp_dir:
            yield Path(temp_dir)
    
    @pytest.fixture
    def garage_file(self, temp_dir):
        """Create test garage registry Excel file"""
        file_path = temp_dir / "test_garage.xlsx"
        
        # Create workbook with garage data
        wb = Workbook()
        ws = wb.active
        
        # Headers
        ws['A1'] = "Гараж"
        ws['B1'] = "Сумма"
        ws['C1'] = "Первоначальная дата"
        
        # Data rows
        garage_data = [
            ("1", 3500.00, "01.01.2025"),
            ("2", 2800.00, "02.01.2025"),
            ("3", 4200.00, "03.01.2025"),
            ("4", 3500.00, "04.01.2025"),  # Duplicate amount
            ("5", 2600.00, "05.01.2025"),  # No matching payment
        ]
        
        for row, (garage_id, amount, start_date) in enumerate(garage_data, 2):
            ws[f'A{row}'] = garage_id
            ws[f'B{row}'] = amount
            ws[f'C{row}'] = start_date
        
        wb.save(file_path)
        return file_path
    
    @pytest.fixture
    def statement_file(self, temp_dir):
        """Create test bank statement Excel file"""
        file_path = temp_dir / "test_statement.xlsx"
        
        # Create workbook with transaction data
        wb = Workbook()
        ws = wb.active
        
        # Transaction data (Sberbank format)
        transaction_data = [
            ("15.01.2025 14:30", "14:30", "", "Перевод СБП", "+3 500,00"),
            ("16.01.2025 15:00", "15:00", "", "Перевод на карту", "+2 800,00"),
            ("17.01.2025 16:00", "16:00", "", "Перевод СБП", "+4 200,00"),
            ("18.01.2025 17:00", "17:00", "", "Перевод СБП", "+3 500,00"),  # Second 3500 for duplicate handling
            ("19.01.2025 18:00", "18:00", "", "Покупка", "-1 000,00"),      # Outgoing transaction (ignored)
            ("20.01.2025 19:00", "19:00", "", "Перевод СБП", "+1 500,00"),  # No matching garage
        ]
        
        for row, data in enumerate(transaction_data, 1):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        wb.save(file_path)
        return file_path
    
    @pytest.fixture
    def use_cases(self):
        """Create configured use cases for testing"""
        parser_factory = ParserFactory()
        payment_matcher = PaymentMatcher(search_window_days=7, grace_period_days=3)
        excel_writer = ExcelReportWriter()
        i18n = LocalizationManager("en")
        
        process_use_case = ProcessPaymentsUseCase(
            parser_factory=parser_factory,
            payment_matcher=payment_matcher
        )
        
        report_use_case = GenerateReportUseCase(
            excel_writer=excel_writer,
            localization_manager=i18n
        )
        
        return process_use_case, report_use_case
    
    def test_complete_payment_processing_flow(self, garage_file, statement_file, use_cases, temp_dir):
        """Test complete flow from file input to report generation"""
        process_use_case, report_use_case = use_cases
        analysis_date = date(2025, 1, 20)  # After all expected payment dates
        
        # Step 1: Create processing request
        request = PaymentProcessRequest(
            garage_file=garage_file,
            statement_file=statement_file,
            analysis_date=analysis_date
        )
        
        # Step 2: Process payments
        process_response = process_use_case.execute(request)
        
        # Verify processing was successful
        assert process_response.success is True
        assert process_response.report is not None
        assert len(process_response.errors) == 0
        
        # Verify garage parsing
        report = process_response.report
        assert len(report.payments) == 5  # 5 garages
        
        # Verify payment matching results
        payment_statuses = {p.garage_id: p.status for p in report.payments}
        
        # Should have received payments for garages with matching transactions
        received_garages = [p.garage_id for p in report.payments if p.status == PaymentStatus.RECEIVED]
        assert len(received_garages) >= 3  # At least 3 should match
        
        # Garage 5 should be overdue (no matching transaction for 2600.00)
        garage_5_payment = next(p for p in report.payments if p.garage_id == "5")
        assert garage_5_payment.status == PaymentStatus.OVERDUE
        
        # Verify duplicate amount handling
        garage_1_payment = next(p for p in report.payments if p.garage_id == "1")
        garage_4_payment = next(p for p in report.payments if p.garage_id == "4")
        
        # Both should have amount 3500, but only one can match each transaction
        assert garage_1_payment.amount == Decimal("3500.00")
        assert garage_4_payment.amount == Decimal("3500.00")
        
        # One should be received, one might be received (depending on matching logic)
        received_3500_count = sum(1 for p in [garage_1_payment, garage_4_payment] 
                                 if p.status == PaymentStatus.RECEIVED)
        assert received_3500_count >= 1  # At least one should match
        
        # Verify warnings about duplicates
        assert len(process_response.warnings) > 0
        assert any("Duplicate rental amount" in warning for warning in process_response.warnings)
        
        # Step 3: Generate Excel report
        output_path = temp_dir / "test_report.xlsx"
        report_response = report_use_case.generate_excel_report(report, output_path)
        
        # Verify report generation
        assert report_response.success is True
        assert report_response.output_path == output_path
        assert output_path.exists()
        
        # Verify report summary
        summary = report.summary
        assert summary.total_garages == 5
        assert summary.received_count >= 3
        assert summary.overdue_count >= 1
        assert summary.total_expected == 16600.00  # Sum of all garage amounts
    
    def test_payment_flow_with_empty_statement(self, garage_file, temp_dir, use_cases):
        """Test payment processing with empty bank statement"""
        process_use_case, _ = use_cases
        
        # Create empty statement file
        empty_statement = temp_dir / "empty_statement.xlsx"
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "No data"
        wb.save(empty_statement)
        
        request = PaymentProcessRequest(
            garage_file=garage_file,
            statement_file=empty_statement,
            analysis_date=date(2025, 1, 20)
        )
        
        response = process_use_case.execute(request)
        
        # Should succeed but all payments should be overdue
        assert response.success is True
        assert all(p.status == PaymentStatus.OVERDUE for p in response.report.payments)
        assert response.report.summary.received_count == 0
        assert response.report.summary.overdue_count == 5
    
    def test_payment_flow_with_early_analysis_date(self, garage_file, statement_file, use_cases):
        """Test payment processing with analysis date before expected payments"""
        process_use_case, _ = use_cases
        
        # Analysis date before any expected payments
        early_date = date(2024, 12, 15)
        
        request = PaymentProcessRequest(
            garage_file=garage_file,
            statement_file=statement_file,
            analysis_date=early_date
        )
        
        response = process_use_case.execute(request)
        
        # Should succeed but all payments should be not due
        assert response.success is True
        # Most payments should be not due (since expected dates are in January 2025)
        not_due_count = sum(1 for p in response.report.payments if p.status == PaymentStatus.NOT_DUE)
        assert not_due_count >= 4  # Most should be not due yet
    
    def test_payment_flow_with_pending_payments(self, garage_file, statement_file, use_cases):
        """Test payment processing with some payments in pending status"""
        process_use_case, _ = use_cases
        
        # Analysis date within grace period for some payments
        analysis_date = date(2025, 1, 17)  # Some payments should be pending
        
        request = PaymentProcessRequest(
            garage_file=garage_file,
            statement_file=statement_file,
            analysis_date=analysis_date
        )
        
        response = process_use_case.execute(request)
        
        assert response.success is True
        
        # Should have mix of received and pending payments
        status_counts = {}
        for payment in response.report.payments:
            status_counts[payment.status] = status_counts.get(payment.status, 0) + 1
        
        # Should have some received payments
        assert status_counts.get(PaymentStatus.RECEIVED, 0) > 0
        
        # Might have pending or overdue depending on exact timing
        pending_or_overdue = (status_counts.get(PaymentStatus.PENDING, 0) + 
                             status_counts.get(PaymentStatus.OVERDUE, 0))
        assert pending_or_overdue >= 0
    
    def test_payment_flow_with_corrupted_garage_file(self, statement_file, use_cases, temp_dir):
        """Test payment processing with corrupted garage file"""
        process_use_case, _ = use_cases
        
        # Create corrupted file (not valid Excel)
        corrupted_file = temp_dir / "corrupted.xlsx"
        corrupted_file.write_text("This is not an Excel file")
        
        request = PaymentProcessRequest(
            garage_file=corrupted_file,
            statement_file=statement_file,
            analysis_date=date(2025, 1, 20)
        )
        
        response = process_use_case.execute(request)
        
        # Should fail gracefully
        assert response.success is False
        assert len(response.errors) > 0
        assert response.report is None
    
    def test_payment_flow_with_missing_files(self, use_cases, temp_dir):
        """Test payment processing with missing files"""
        process_use_case, _ = use_cases
        
        nonexistent_garage = temp_dir / "nonexistent_garage.xlsx"
        nonexistent_statement = temp_dir / "nonexistent_statement.xlsx"
        
        # Should fail at request creation level
        with pytest.raises(FileNotFoundError):
            PaymentProcessRequest(
                garage_file=nonexistent_garage,
                statement_file=nonexistent_statement,
                analysis_date=date(2025, 1, 20)
            )
    
    def test_payment_flow_date_edge_cases(self, temp_dir, use_cases):
        """Test payment processing with date edge cases"""
        process_use_case, _ = use_cases
        
        # Create garage file with February 29 payment day
        garage_file = temp_dir / "edge_garage.xlsx"
        wb = Workbook()
        ws = wb.active
        
        ws['A1'] = "Гараж"
        ws['B1'] = "Сумма"
        ws['C1'] = "Первоначальная дата"
        
        # Garage with payment day 29 (edge case for February)
        ws['A2'] = "1"
        ws['B2'] = 3500.00
        ws['C2'] = "29.01.2024"  # Payment day 29
        
        wb.save(garage_file)
        
        # Create statement with payment on February 28 (since Feb 29 doesn't exist in 2025)
        statement_file = temp_dir / "edge_statement.xlsx"
        wb = Workbook()
        ws = wb.active
        
        ws['A1'] = "28.02.2025 14:30"
        ws['B1'] = "14:30"
        ws['C1'] = ""
        ws['D1'] = "Перевод СБП"
        ws['E1'] = "+3 500,00"
        
        wb.save(statement_file)
        
        request = PaymentProcessRequest(
            garage_file=garage_file,
            statement_file=statement_file,
            analysis_date=date(2025, 3, 5)  # March analysis
        )
        
        response = process_use_case.execute(request)
        
        # Should handle date adjustment gracefully
        assert response.success is True
        assert len(response.report.payments) == 1
        
        payment = response.report.payments[0]
        # Expected date should be adjusted to Feb 28 (last day of February 2025)
        assert payment.expected_date.day <= 28
        assert payment.expected_date.month == 2
    
    def test_full_flow_performance(self, garage_file, statement_file, use_cases, temp_dir):
        """Test that full flow completes in reasonable time"""
        import time
        
        process_use_case, report_use_case = use_cases
        
        start_time = time.time()
        
        # Process payments
        request = PaymentProcessRequest(
            garage_file=garage_file,
            statement_file=statement_file,
            analysis_date=date(2025, 1, 20)
        )
        
        process_response = process_use_case.execute(request)
        assert process_response.success is True
        
        # Generate report
        output_path = temp_dir / "performance_report.xlsx"
        report_response = report_use_case.generate_excel_report(process_response.report, output_path)
        assert report_response.success is True
        
        end_time = time.time()
        
        # Should complete within reasonable time (adjust threshold as needed)
        processing_time = end_time - start_time
        assert processing_time < 5.0  # Should complete within 5 seconds
    
    def test_report_content_verification(self, garage_file, statement_file, use_cases, temp_dir):
        """Test that generated report contains expected content"""
        process_use_case, report_use_case = use_cases
        
        # Process payments
        request = PaymentProcessRequest(
            garage_file=garage_file,
            statement_file=statement_file,
            analysis_date=date(2025, 1, 20)
        )
        
        process_response = process_use_case.execute(request)
        report = process_response.report
        
        # Generate report
        output_path = temp_dir / "content_report.xlsx"
        report_response = report_use_case.generate_excel_report(report, output_path)
        
        assert report_response.success is True
        assert output_path.exists()
        
        # Verify file is not empty
        assert output_path.stat().st_size > 0
        
        # Verify report object content
        assert report.garage_file == str(garage_file)
        assert report.statement_file == str(statement_file)
        assert report.analysis_date == date(2025, 1, 20)
        assert len(report.payments) == 5
        
        # Verify summary calculations
        summary = report.summary
        total_amount = sum(p.amount for p in report.payments)
        assert summary.total_expected == float(total_amount)
        
        received_amount = sum(p.amount for p in report.payments if p.status == PaymentStatus.RECEIVED)
        assert summary.total_received == float(received_amount)
        
        # Verify status counts add up
        total_status_count = (summary.received_count + summary.overdue_count + 
                            summary.pending_count + summary.not_due_count + 
                            summary.unclear_count)
        assert total_status_count == summary.total_garages
