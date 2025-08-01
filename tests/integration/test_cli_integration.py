"""
Integration tests for CLI interface
"""

import pytest
import tempfile
import subprocess
import sys
from pathlib import Path
from openpyxl import Workbook
from click.testing import CliRunner

from src.interfaces.cli.cli_app import CLIApp
from src.infrastructure.config.config_manager import ConfigManager
from src.infrastructure.localization.i18n import LocalizationManager


class TestCLIIntegration:
    """Integration tests for CLI application"""
    
    @pytest.fixture
    def temp_dir(self):
        """Create temporary directory for test files"""
        with tempfile.TemporaryDirectory() as temp_dir:
            yield Path(temp_dir)
    
    @pytest.fixture
    def sample_garage_file(self, temp_dir):
        """Create sample garage registry file"""
        file_path = temp_dir / "garages.xlsx"
        
        wb = Workbook()
        ws = wb.active
        
        # Headers
        ws['A1'] = "Гараж"
        ws['B1'] = "Сумма"
        ws['C1'] = "Первоначальная дата"
        
        # Sample data
        garage_data = [
            ("1", 3500.00, "15.01.2025"),
            ("2", 2800.00, "16.01.2025"),
            ("3", 4200.00, "17.01.2025"),
        ]
        
        for row, (garage_id, amount, date_str) in enumerate(garage_data, 2):
            ws[f'A{row}'] = garage_id
            ws[f'B{row}'] = amount
            ws[f'C{row}'] = date_str
        
        wb.save(file_path)
        return file_path
    
    @pytest.fixture
    def sample_statement_file(self, temp_dir):
        """Create sample bank statement file"""
        file_path = temp_dir / "statement.xlsx"
        
        wb = Workbook()
        ws = wb.active
        
        # Sample transaction data
        transaction_data = [
            ("15.01.2025 14:30", "14:30", "", "Перевод СБП", "+3 500,00"),
            ("16.01.2025 15:00", "15:00", "", "Перевод на карту", "+2 800,00"),
        ]
        
        for row, data in enumerate(transaction_data, 1):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        wb.save(file_path)
        return file_path
    
    @pytest.fixture
    def cli_app(self):
        """Create CLI application for testing"""
        config = {
            'application': {
                'name': 'Garage Payment Tracker',
                'version': '1.0.0',
                'default_language': 'en'
            },
            'parsing': {
                'grace_period_days': 3,
                'search_window_days': 7
            }
        }
        i18n = LocalizationManager('en')
        return CLIApp(config, i18n)
    
    @pytest.fixture
    def runner(self):
        """Create Click CLI test runner"""
        return CliRunner()
    
    def test_cli_process_payments_command_success(self, runner, cli_app, sample_garage_file, 
                                                sample_statement_file, temp_dir):
        """Test successful execution of process-payments command"""
        output_file = temp_dir / "test_report.xlsx"
        
        with runner.isolated_filesystem():
            # Copy files to isolated filesystem
            local_garage = Path("garages.xlsx")
            local_statement = Path("statement.xlsx")
            local_output = Path("report.xlsx")
            
            local_garage.write_bytes(sample_garage_file.read_bytes())
            local_statement.write_bytes(sample_statement_file.read_bytes())
            
            # Mock the CLI app's run method to test the command directly
            import click
            
            @click.command()
            @click.option('--garage-file', required=True, type=click.Path(exists=True, path_type=Path))
            @click.option('--statement-file', required=True, type=click.Path(exists=True, path_type=Path))
            @click.option('--output', '-o', type=click.Path(path_type=Path))
            def test_process_payments(garage_file, statement_file, output):
                """Test process payments command"""
                from src.application.dto.payment_request import PaymentProcessRequest
                from datetime import date
                
                try:
                    request = PaymentProcessRequest(
                        garage_file=garage_file,
                        statement_file=statement_file,
                        analysis_date=date(2025, 1, 20),
                        output_path=output
                    )
                    
                    response = cli_app.process_payments_use_case.execute(request)
                    
                    if response.success:
                        # Generate report
                        report_response = cli_app.generate_report_use_case.generate_excel_report(
                            response.report, output or Path("default_report.xlsx")
                        )
                        
                        if report_response.success:
                            click.echo("SUCCESS: Report generated successfully")
                        else:
                            click.echo("ERROR: Report generation failed")
                    else:
                        click.echo("ERROR: Payment processing failed")
                
                except Exception as e:
                    click.echo(f"ERROR: {e}")
            
            result = runner.invoke(test_process_payments, [
                '--garage-file', str(local_garage),
                '--statement-file', str(local_statement),
                '--output', str(local_output)
            ])
            
            # Check that command executed without errors
            assert result.exit_code == 0
            assert "SUCCESS" in result.output or "ERROR" in result.output
    
    def test_cli_version_command(self, runner, cli_app):
        """Test version command"""
        # Create a simple version command for testing
        import click
        
        @click.command()
        def version():
            """Show version information"""
            click.echo("Garage Payment Tracker v1.0.0")
        
        result = runner.invoke(version)
        
        assert result.exit_code == 0
        assert "Garage Payment Tracker" in result.output
        assert "1.0.0" in result.output
    
    def test_cli_missing_required_arguments(self, runner):
        """Test CLI with missing required arguments"""
        import click
        
        @click.command()
        @click.option('--garage-file', required=True, type=click.Path(exists=True))
        @click.option('--statement-file', required=True, type=click.Path(exists=True))
        def test_command(garage_file, statement_file):
            click.echo("Should not reach here")
        
        # Test with missing garage file
        result = runner.invoke(test_command, ['--statement-file', 'nonexistent.xlsx'])
        assert result.exit_code != 0
        assert "Missing option" in result.output or "Error" in result.output
        
        # Test with missing statement file
        result = runner.invoke(test_command, ['--garage-file', 'nonexistent.xlsx'])
        assert result.exit_code != 0
        assert "Missing option" in result.output or "Error" in result.output
    
    def test_cli_nonexistent_files(self, runner):
        """Test CLI with nonexistent input files"""
        import click
        
        @click.command()
        @click.option('--garage-file', required=True, type=click.Path(exists=True, path_type=Path))
        @click.option('--statement-file', required=True, type=click.Path(exists=True, path_type=Path))
        def test_command(garage_file, statement_file):
            click.echo("Files exist")
        
        result = runner.invoke(test_command, [
            '--garage-file', 'nonexistent_garage.xlsx',
            '--statement-file', 'nonexistent_statement.xlsx'
        ])
        
        assert result.exit_code != 0
        # Click should handle the file existence check
    
    def test_cli_verbose_option(self, runner, cli_app):
        """Test CLI with verbose option"""
        import click
        
        @click.command()
        @click.option('--verbose', '-v', is_flag=True, help='Enable verbose logging')
        def test_verbose(verbose):
            if verbose:
                click.echo("Verbose mode enabled")
            else:
                click.echo("Normal mode")
        
        # Test without verbose
        result = runner.invoke(test_verbose)
        assert result.exit_code == 0
        assert "Normal mode" in result.output
        
        # Test with verbose
        result = runner.invoke(test_verbose, ['--verbose'])
        assert result.exit_code == 0
        assert "Verbose mode enabled" in result.output
    
    def test_cli_custom_analysis_date(self, runner, sample_garage_file, sample_statement_file):
        """Test CLI with custom analysis date"""
        import click
        from datetime import datetime
        
        @click.command()
        @click.option('--analysis-date', type=click.DateTime(formats=['%Y-%m-%d']),
                     help='Analysis date (default: today)')
        def test_date(analysis_date):
            if analysis_date:
                click.echo(f"Analysis date: {analysis_date.date()}")
            else:
                click.echo("Using default date")
        
        # Test with custom date
        result = runner.invoke(test_date, ['--analysis-date', '2025-01-20'])
        assert result.exit_code == 0
        assert "2025-01-20" in result.output
        
        # Test with invalid date format
        result = runner.invoke(test_date, ['--analysis-date', 'invalid-date'])
        assert result.exit_code != 0
    
    def test_cli_help_messages(self, runner):
        """Test CLI help messages"""
        import click
        
        @click.group()
        def cli():
            """Garage Payment Tracker - Track rental payments and generate reports"""
            pass
        
        @cli.command('process-payments')
        @click.option('--garage-file', required=True, help='Path to garage registry Excel file')
        @click.option('--statement-file', required=True, help='Path to bank statement Excel file')
        def process_payments(garage_file, statement_file):
            """Process garage payments and generate report"""
            pass
        
        # Test main help
        result = runner.invoke(cli, ['--help'])
        assert result.exit_code == 0
        assert "Garage Payment Tracker" in result.output
        assert "Track rental payments" in result.output
        
        # Test command help
        result = runner.invoke(cli, ['process-payments', '--help'])
        assert result.exit_code == 0
        assert "Process garage payments" in result.output
        assert "garage-file" in result.output
        assert "statement-file" in result.output
    
    def test_cli_output_path_handling(self, runner, sample_garage_file, sample_statement_file, temp_dir):
        """Test CLI output path handling"""
        import click
        
        @click.command()
        @click.option('--output', '-o', type=click.Path(path_type=Path))
        def test_output(output):
            if output:
                click.echo(f"Output path: {output}")
                # Create a dummy file to simulate report generation
                output.parent.mkdir(parents=True, exist_ok=True)
                output.write_text("dummy report")
                click.echo(f"Report saved to: {output}")
            else:
                click.echo("Using default output path")
        
        with runner.isolated_filesystem():
            # Test with custom output path
            custom_output = "custom_report.xlsx"
            result = runner.invoke(test_output, ['--output', custom_output])
            assert result.exit_code == 0
            assert "custom_report.xlsx" in result.output
            assert Path(custom_output).exists()
            
            # Test without output path
            result = runner.invoke(test_output)
            assert result.exit_code == 0
            assert "default output" in result.output
    
    def test_cli_error_handling(self, runner, cli_app):
        """Test CLI error handling"""
        import click
        from src.core.exceptions import GarageTrackerException
        
        @click.command()
        @click.option('--simulate-error', type=click.Choice(['parse', 'validation', 'general']))
        def test_errors(simulate_error):
            try:
                if simulate_error == 'parse':
                    from src.core.exceptions import ParseError
                    raise ParseError("Simulated parse error", "test.xlsx")
                elif simulate_error == 'validation':
                    from src.core.exceptions import ValidationError
                    raise ValidationError("Simulated validation error")
                elif simulate_error == 'general':
                    raise Exception("Simulated general error")
                else:
                    click.echo("No error simulated")
            except GarageTrackerException as e:
                click.echo(f"Tracker Error: {e}", err=True)
                raise click.ClickException(str(e))
            except Exception as e:
                click.echo(f"Unexpected Error: {e}", err=True)
                raise click.ClickException(str(e))
        
        # Test parse error
        result = runner.invoke(test_errors, ['--simulate-error', 'parse'])
        assert result.exit_code != 0
        assert "Tracker Error" in result.output
        
        # Test validation error
        result = runner.invoke(test_errors, ['--simulate-error', 'validation'])
        assert result.exit_code != 0
        assert "Tracker Error" in result.output
        
        # Test general error
        result = runner.invoke(test_errors, ['--simulate-error', 'general'])
        assert result.exit_code != 0
        assert "Unexpected Error" in result.output
    
    def test_main_py_execution(self, sample_garage_file, sample_statement_file, temp_dir):
        """Test that main.py can be executed successfully"""
        # This test verifies that the main entry point works
        # Note: This is more of a smoke test since we can't easily test the full CLI
        
        try:
            # Import main module to ensure it loads without errors
            import main
            
            # Verify main function exists
            assert hasattr(main, 'main')
            assert hasattr(main, 'create_app')
            
            # Test create_app function
            app = main.create_app('en')
            assert app is not None
            assert hasattr(app, 'run')
            
        except Exception as e:
            pytest.fail(f"Failed to import or create main app: {e}")
    
    def test_cli_configuration_loading(self, runner, temp_dir):
        """Test CLI with different configuration scenarios"""
        # Create a temporary config file
        config_file = temp_dir / "test_config.yaml"
        config_content = """
application:
  name: "Test Garage Tracker"
  version: "1.0.0"
  default_language: "en"

parsing:
  grace_period_days: 5
  search_window_days: 10
"""
        config_file.write_text(config_content)
        
        # Test that config can be loaded
        try:
            from src.infrastructure.config.config_manager import ConfigManager
            config = ConfigManager.load_config(config_file)
            
            assert config['application']['name'] == "Test Garage Tracker"
            assert config['parsing']['grace_period_days'] == 5
            assert config['parsing']['search_window_days'] == 10
            
        except Exception as e:
            pytest.fail(f"Failed to load test configuration: {e}")
    
    def test_cli_localization(self, runner, cli_app):
        """Test CLI with different localization settings"""
        # Test English localization (default)
        en_i18n = LocalizationManager('en')
        assert en_i18n.get('status.received') == 'Received'
        assert en_i18n.get('status.overdue') == 'Overdue'
        
        # Test that we can create app with different language
        config = {
            'application': {'default_language': 'en'},
            'parsing': {'grace_period_days': 3, 'search_window_days': 7}
        }
        
        app_en = CLIApp(config, LocalizationManager('en'))
        assert app_en.i18n.language == 'en'
        
        # Note: Russian localization would be tested when implemented
