"""
Tests for payment period functionality
"""

import unittest
from datetime import date
from pathlib import Path
from tempfile import NamedTemporaryFile

from openpyxl import Workbook

import sys
sys.path.insert(0, 'src')

from src.core.models.payment_period import PaymentPeriod
from src.parsers.file_parsers.sberbank_parser import SberbankStatementParser
from src.core.exceptions import ParseError


class TestPaymentPeriod(unittest.TestCase):
    """Test payment period model"""
    
    def test_payment_period_creation(self):
        """Test PaymentPeriod creation and validation"""
        period = PaymentPeriod(
            start_date=date(2025, 5, 1),
            end_date=date(2025, 6, 12),
            source_text="Итого по операциям с 01.05.2025 по 12.06.2025"
        )
        
        self.assertEqual(period.start_date, date(2025, 5, 1))
        self.assertEqual(period.end_date, date(2025, 6, 12))
        self.assertEqual(period.duration_days, 43)
        self.assertEqual(period.target_month, date(2025, 5, 1))
    
    def test_payment_period_invalid_dates(self):
        """Test PaymentPeriod validation with invalid dates"""
        with self.assertRaises(ValueError):
            PaymentPeriod(
                start_date=date(2025, 6, 12),
                end_date=date(2025, 5, 1),  # End before start
                source_text="Invalid period"
            )
    
    def test_contains_date(self):
        """Test date containment check"""
        period = PaymentPeriod(
            start_date=date(2025, 5, 1),
            end_date=date(2025, 6, 12),
            source_text="Test period"
        )
        
        self.assertTrue(period.contains_date(date(2025, 5, 15)))
        self.assertTrue(period.contains_date(date(2025, 5, 1)))  # Start date
        self.assertTrue(period.contains_date(date(2025, 6, 12)))  # End date
        self.assertFalse(period.contains_date(date(2025, 4, 30)))  # Before start
        self.assertFalse(period.contains_date(date(2025, 6, 13)))  # After end


class TestSberbankPeriodExtraction(unittest.TestCase):
    """Test period extraction from Sberbank statements"""
    
    def setUp(self):
        self.parser = SberbankStatementParser()
    
    def test_extract_period_from_text(self):
        """Test period extraction from Excel file with period text"""
        # Create temporary Excel file with period text
        with NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            workbook = Workbook()
            worksheet = workbook.active
            
            # Add some dummy data
            worksheet['A1'] = '01.05.2025 10:30'
            worksheet['B1'] = 'Перевод на карту'
            worksheet['C1'] = '+3500,00'
            
            # Add period summary line
            worksheet['A5'] = 'Итого по операциям с 01.05.2025 по 12.06.2025'
            
            workbook.save(tmp_file.name)
            workbook.close()
            
            tmp_path = Path(tmp_file.name)
            
            try:
                period = self.parser.extract_payment_period(tmp_path)
                
                self.assertIsNotNone(period)
                self.assertEqual(period.start_date, date(2025, 5, 1))
                self.assertEqual(period.end_date, date(2025, 6, 12))
                self.assertIn("01.05.2025", period.source_text)
                self.assertIn("12.06.2025", period.source_text)
                
            finally:
                tmp_path.unlink()  # Clean up
    
    def test_extract_period_not_found(self):
        """Test period extraction when no period text exists"""
        # Create temporary Excel file without period text
        with NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            workbook = Workbook()
            worksheet = workbook.active
            
            # Add some dummy data without period summary
            worksheet['A1'] = '01.05.2025 10:30'
            worksheet['B1'] = 'Перевод на карту'
            worksheet['C1'] = '+3500,00'
            
            workbook.save(tmp_file.name)
            workbook.close()
            
            tmp_path = Path(tmp_file.name)
            
            try:
                period = self.parser.extract_payment_period(tmp_path)
                self.assertIsNone(period)
                
            finally:
                tmp_path.unlink()  # Clean up
    
    def test_extract_period_invalid_file(self):
        """Test period extraction with invalid file"""
        invalid_path = Path('nonexistent_file.xlsx')
        period = self.parser.extract_payment_period(invalid_path)
        self.assertIsNone(period)
    
    def test_extract_period_case_insensitive(self):
        """Test period extraction is case insensitive"""
        # Create temporary Excel file with uppercase period text
        with NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            workbook = Workbook()
            worksheet = workbook.active
            
            # Add some dummy transaction data first to make it look like Sberbank file
            worksheet['A1'] = '01.05.2025 10:30'
            worksheet['B1'] = 'Перевод на карту'
            worksheet['C1'] = '+3500,00'
            
            # Add period summary line in uppercase
            worksheet['A5'] = 'ИТОГО ПО ОПЕРАЦИЯМ С 01.05.2025 ПО 12.06.2025'
            
            workbook.save(tmp_file.name)
            workbook.close()
            
            tmp_path = Path(tmp_file.name)
            
            try:
                period = self.parser.extract_payment_period(tmp_path)
                
                self.assertIsNotNone(period)
                self.assertEqual(period.start_date, date(2025, 5, 1))
                self.assertEqual(period.end_date, date(2025, 6, 12))
                
            finally:
                tmp_path.unlink()  # Clean up


if __name__ == '__main__':
    unittest.main()