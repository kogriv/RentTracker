"""
Unit tests for SberbankStatementParser
"""

import pytest
from pathlib import Path
from decimal import Decimal
from datetime import date, datetime
from unittest.mock import Mock, patch

from src.parsers.file_parsers.sberbank_parser import SberbankStatementParser
from src.core.exceptions import ParseError


class TestSberbankStatementParser:
    """Test cases for SberbankStatementParser"""
    
    @pytest.fixture
    def sberbank_parser(self):
        """Create SberbankStatementParser instance for testing"""
        return SberbankStatementParser()
    
    def test_get_supported_formats(self, sberbank_parser):
        """Test supported file formats"""
        formats = sberbank_parser.get_supported_formats()
        assert '.xlsx' in formats
        assert '.xls' in formats
    
    def test_extract_date_from_datetime_cell(self, sberbank_parser):
        """Test date extraction from datetime cell"""
        mock_row = [
            Mock(value=datetime(2025, 1, 15, 14, 30)),
            Mock(value=None)
        ]
        
        extracted_date = sberbank_parser._extract_date(mock_row)
        assert extracted_date == date(2025, 1, 15)
    
    def test_extract_date_from_string_cell(self, sberbank_parser):
        """Test date extraction from string cell"""
        mock_row = [
            Mock(value="15.01.2025 14:30"),
            Mock(value=None)
        ]
        
        extracted_date = sberbank_parser._extract_date(mock_row)
        assert extracted_date == date(2025, 1, 15)
    
    def test_extract_date_from_multiple_columns(self, sberbank_parser):
        """Test date extraction from multiple columns"""
        mock_row = [
            Mock(value=None),
            Mock(value="15.01.2025"),
            Mock(value=None)
        ]
        
        extracted_date = sberbank_parser._extract_date(mock_row)
        assert extracted_date == date(2025, 1, 15)
    
    def test_extract_date_no_date_found(self, sberbank_parser):
        """Test date extraction when no date found"""
        mock_row = [
            Mock(value="Some text"),
            Mock(value=123.45),
            Mock(value=None)
        ]
        
        extracted_date = sberbank_parser._extract_date(mock_row)
        assert extracted_date is None
    
    def test_extract_amount_standard_format(self, sberbank_parser):
        """Test amount extraction from standard format"""
        mock_row = [
            Mock(value="15.01.2025"),
            Mock(value=None),
            Mock(value=None),
            Mock(value="Перевод"),
            Mock(value="+3 500,00")
        ]
        
        extracted_amount = sberbank_parser._extract_amount(mock_row)
        assert extracted_amount == Decimal("3500.00")
    
    def test_extract_amount_with_spaces(self, sberbank_parser):
        """Test amount extraction with spaces"""
        mock_row = [
            Mock(value=None),
            Mock(value=None),
            Mock(value=None),
            Mock(value=None),
            Mock(value="+ 2 800,50")
        ]
        
        extracted_amount = sberbank_parser._extract_amount(mock_row)
        assert extracted_amount == Decimal("2800.50")
    
    def test_extract_amount_simple_format(self, sberbank_parser):
        """Test amount extraction from simple format"""
        mock_row = [
            Mock(value=None),
            Mock(value=None),
            Mock(value="+4200.00"),
            Mock(value=None),
            Mock(value=None)
        ]
        
        extracted_amount = sberbank_parser._extract_amount(mock_row)
        assert extracted_amount == Decimal("4200.00")
    
    def test_extract_amount_no_plus_sign(self, sberbank_parser):
        """Test amount extraction without plus sign (should return None)"""
        mock_row = [
            Mock(value="15.01.2025"),
            Mock(value=None),
            Mock(value=None),
            Mock(value="Перевод"),
            Mock(value="3500,00")  # No plus sign
        ]
        
        extracted_amount = sberbank_parser._extract_amount(mock_row)
        assert extracted_amount is None
    
    def test_extract_amount_no_amount_found(self, sberbank_parser):
        """Test amount extraction when no amount found"""
        mock_row = [
            Mock(value="15.01.2025"),
            Mock(value="Text"),
            Mock(value="More text"),
            Mock(value="Category"),
            Mock(value="No amount here")
        ]
        
        extracted_amount = sberbank_parser._extract_amount(mock_row)
        assert extracted_amount is None
    
    def test_extract_category_standard(self, sberbank_parser):
        """Test category extraction"""
        mock_row = [
            Mock(value="15.01.2025"),
            Mock(value=None),
            Mock(value=None),
            Mock(value="Перевод СБП"),
            Mock(value="+3500,00")
        ]
        
        extracted_category = sberbank_parser._extract_category(mock_row)
        assert extracted_category == "Перевод СБП"
    
    def test_extract_category_from_different_column(self, sberbank_parser):
        """Test category extraction from different column"""
        mock_row = [
            Mock(value="15.01.2025"),
            Mock(value="Перевод на карту"),
            Mock(value=None),
            Mock(value="Other text"),
            Mock(value="+3500,00")
        ]
        
        extracted_category = sberbank_parser._extract_category(mock_row)
        assert extracted_category == "Перевод на карту"
    
    def test_extract_category_no_keywords(self, sberbank_parser):
        """Test category extraction with no transfer keywords"""
        mock_row = [
            Mock(value="15.01.2025"),
            Mock(value="Some text"),
            Mock(value="Other text"),
            Mock(value="Random category"),
            Mock(value="+3500,00")
        ]
        
        extracted_category = sberbank_parser._extract_category(mock_row)
        assert extracted_category == "Unknown Transfer"
    
    def test_is_transfer_category_valid_keywords(self, sberbank_parser):
        """Test transfer category validation with valid keywords"""
        valid_categories = [
            "Перевод СБП",
            "Перевод на карту",
            "Transfer",
            "Mobile transfer"
        ]
        
        for category in valid_categories:
            assert sberbank_parser._is_transfer_category(category) is True
    
    def test_is_transfer_category_invalid_keywords(self, sberbank_parser):
        """Test transfer category validation with invalid keywords"""
        invalid_categories = [
            "Покупка",
            "Оплата услуг",
            "Снятие наличных",
            "Random text",
            None,
            ""
        ]
        
        for category in invalid_categories:
            assert sberbank_parser._is_transfer_category(category) is False
    
    def test_parse_transaction_row_valid(self, sberbank_parser):
        """Test parsing valid transaction row"""
        mock_row = [
            Mock(value="15.01.2025 14:30"),
            Mock(value="14:30"),
            Mock(value=None),
            Mock(value="Перевод СБП"),
            Mock(value="+3 500,00")
        ]
        
        transaction = sberbank_parser._parse_transaction_row(mock_row, 5)
        
        assert transaction is not None
        assert transaction.date == date(2025, 1, 15)
        assert transaction.amount == Decimal("3500.00")
        assert transaction.category == "Перевод СБП"
        assert "row_5" in transaction.source
    
    def test_parse_transaction_row_missing_date(self, sberbank_parser):
        """Test parsing transaction row with missing date"""
        mock_row = [
            Mock(value=None),
            Mock(value=None),
            Mock(value=None),
            Mock(value="Перевод СБП"),
            Mock(value="+3500,00")
        ]
        
        transaction = sberbank_parser._parse_transaction_row(mock_row, 5)
        assert transaction is None
    
    def test_parse_transaction_row_missing_amount(self, sberbank_parser):
        """Test parsing transaction row with missing amount"""
        mock_row = [
            Mock(value="15.01.2025"),
            Mock(value=None),
            Mock(value=None),
            Mock(value="Перевод СБП"),
            Mock(value="No amount")
        ]
        
        transaction = sberbank_parser._parse_transaction_row(mock_row, 5)
        assert transaction is None
    
    def test_parse_transaction_row_negative_amount(self, sberbank_parser):
        """Test parsing transaction row with negative amount (outgoing)"""
        mock_row = [
            Mock(value="15.01.2025"),
            Mock(value=None),
            Mock(value=None),
            Mock(value="Перевод СБП"),
            Mock(value="-3500,00")  # Negative amount
        ]
        
        # This should return None because we only want incoming transactions
        transaction = sberbank_parser._parse_transaction_row(mock_row, 5)
        assert transaction is None
    
    def test_parse_transaction_row_invalid_category(self, sberbank_parser):
        """Test parsing transaction row with invalid category"""
        mock_row = [
            Mock(value="15.01.2025"),
            Mock(value=None),
            Mock(value=None),
            Mock(value="Покупка"),  # Not a transfer
            Mock(value="+3500,00")
        ]
        
        transaction = sberbank_parser._parse_transaction_row(mock_row, 5)
        assert transaction is None
    
    def test_looks_like_sberbank_row_valid(self, sberbank_parser):
        """Test Sberbank row detection with valid patterns"""
        # Row with date pattern
        mock_row = [Mock(value="15.01.2025 14:30"), Mock(value="Text")]
        assert sberbank_parser._looks_like_sberbank_row(mock_row) is True
        
        # Row with amount pattern
        mock_row = [Mock(value="Text"), Mock(value="+3500,00")]
        assert sberbank_parser._looks_like_sberbank_row(mock_row) is True
    
    def test_looks_like_sberbank_row_invalid(self, sberbank_parser):
        """Test Sberbank row detection with invalid patterns"""
        # Row without patterns
        mock_row = [Mock(value="Random text"), Mock(value="More text")]
        assert sberbank_parser._looks_like_sberbank_row(mock_row) is False
        
        # Empty row
        mock_row = []
        assert sberbank_parser._looks_like_sberbank_row(mock_row) is False
        
        # Row with single cell
        mock_row = [Mock(value="Text")]
        assert sberbank_parser._looks_like_sberbank_row(mock_row) is False
    
    @patch('src.parsers.file_parsers.sberbank_parser.load_workbook')
    def test_validate_source_valid_sberbank_file(self, mock_load_workbook, sberbank_parser):
        """Test source validation for valid Sberbank file"""
        mock_workbook = Mock()
        mock_worksheet = Mock()
        mock_workbook.active = mock_worksheet
        mock_load_workbook.return_value = mock_workbook
        
        # Mock rows that look like Sberbank data
        mock_rows = [
            [Mock(value="15.01.2025 14:30"), Mock(value="Text")],
            [Mock(value="16.01.2025 15:00"), Mock(value="+3500,00")],
        ]
        mock_worksheet.iter_rows.return_value = mock_rows
        
        test_file = Path("test.xlsx")
        
        with patch.object(test_file, 'exists', return_value=True):
            with patch.object(test_file, 'suffix', '.xlsx'):
                result = sberbank_parser.validate_source(test_file)
                assert result is True
        
        mock_workbook.close.assert_called_once()
    
    @patch('src.parsers.file_parsers.sberbank_parser.load_workbook')
    def test_validate_source_non_sberbank_file(self, mock_load_workbook, sberbank_parser):
        """Test source validation for non-Sberbank Excel file"""
        mock_workbook = Mock()
        mock_worksheet = Mock()
        mock_workbook.active = mock_worksheet
        mock_load_workbook.return_value = mock_workbook
        
        # Mock rows that don't look like Sberbank data
        mock_rows = [
            [Mock(value="Random text"), Mock(value="More text")],
            [Mock(value="No patterns"), Mock(value="Here either")],
        ]
        mock_worksheet.iter_rows.return_value = mock_rows
        
        test_file = Path("test.xlsx")
        
        with patch.object(test_file, 'exists', return_value=True):
            with patch.object(test_file, 'suffix', '.xlsx'):
                result = sberbank_parser.validate_source(test_file)
                assert result is False
        
        mock_workbook.close.assert_called_once()
    
    @patch('src.parsers.file_parsers.sberbank_parser.load_workbook')
    def test_parse_transactions_successful(self, mock_load_workbook, sberbank_parser):
        """Test successful transaction parsing"""
        mock_workbook = Mock()
        mock_worksheet = Mock()
        mock_workbook.active = mock_worksheet
        mock_load_workbook.return_value = mock_workbook
        
        # Mock rows with valid transaction data
        mock_rows = [
            [Mock(value="15.01.2025 14:30"), Mock(value="14:30"), Mock(value=None), 
             Mock(value="Перевод СБП"), Mock(value="+3 500,00")],
            [Mock(value="16.01.2025 15:00"), Mock(value="15:00"), Mock(value=None),
             Mock(value="Перевод на карту"), Mock(value="+2 800,50")],
            [Mock(value="17.01.2025 16:00"), Mock(value="16:00"), Mock(value=None),
             Mock(value="Покупка"), Mock(value="-1000,00")],  # Outgoing - should be filtered
        ]
        mock_worksheet.iter_rows.return_value = mock_rows
        
        test_file = Path("test.xlsx")
        
        with patch.object(sberbank_parser, 'validate_source', return_value=True):
            transactions = sberbank_parser.parse_transactions(test_file)
        
        # Should only get 2 incoming transactions (third is outgoing)
        assert len(transactions) == 2
        assert transactions[0].amount == Decimal("3500.00")
        assert transactions[1].amount == Decimal("2800.50")
        
        mock_workbook.close.assert_called_once()
    
    @patch('src.parsers.file_parsers.sberbank_parser.load_workbook')
    def test_parse_transactions_empty_file(self, mock_load_workbook, sberbank_parser):
        """Test transaction parsing with empty file"""
        mock_workbook = Mock()
        mock_worksheet = Mock()
        mock_workbook.active = mock_worksheet
        mock_load_workbook.return_value = mock_workbook
        
        # Mock empty rows
        mock_rows = [
            [Mock(value=None), Mock(value=None), Mock(value=None)],
        ]
        mock_worksheet.iter_rows.return_value = mock_rows
        
        test_file = Path("test.xlsx")
        
        with patch.object(sberbank_parser, 'validate_source', return_value=True):
            transactions = sberbank_parser.parse_transactions(test_file)
        
        assert len(transactions) == 0
    
    def test_parse_transactions_invalid_file(self, sberbank_parser):
        """Test transaction parsing with invalid file"""
        test_file = Path("invalid.xlsx")
        
        with patch.object(sberbank_parser, 'validate_source', return_value=False):
            with pytest.raises(ParseError, match="Invalid Excel file"):
                sberbank_parser.parse_transactions(test_file)
    
    @patch('src.parsers.file_parsers.sberbank_parser.load_workbook')
    def test_parse_transactions_workbook_error(self, mock_load_workbook, sberbank_parser):
        """Test transaction parsing with workbook loading error"""
        from openpyxl.utils.exceptions import InvalidFileException
        mock_load_workbook.side_effect = InvalidFileException("Invalid file")
        
        test_file = Path("test.xlsx")
        
        with patch.object(sberbank_parser, 'validate_source', return_value=True):
            with pytest.raises(ParseError, match="Invalid Excel file format"):
                sberbank_parser.parse_transactions(test_file)
    
    def test_amount_pattern_matching(self, sberbank_parser):
        """Test various amount pattern matching scenarios"""
        amount_test_cases = [
            ("+3 500,00", Decimal("3500.00")),
            ("+ 2 800,50", Decimal("2800.50")),
            ("+4200.00", Decimal("4200.00")),
            ("+1,234.56", Decimal("1234.56")),
            ("+ 10 000,00", Decimal("10000.00")),
        ]
        
        for amount_text, expected_amount in amount_test_cases:
            mock_row = [Mock(value=amount_text)]
            extracted_amount = sberbank_parser._extract_amount(mock_row)
            assert extracted_amount == expected_amount, f"Failed for amount: {amount_text}"
    
    def test_date_pattern_matching(self, sberbank_parser):
        """Test various date pattern matching scenarios"""
        date_test_cases = [
            ("15.01.2025", date(2025, 1, 15)),
            ("01.12.2024 14:30", date(2024, 12, 1)),
            ("31.03.2025 23:59", date(2025, 3, 31)),
        ]
        
        for date_text, expected_date in date_test_cases:
            mock_row = [Mock(value=date_text)]
            extracted_date = sberbank_parser._extract_date(mock_row)
            assert extracted_date == expected_date, f"Failed for date: {date_text}"
