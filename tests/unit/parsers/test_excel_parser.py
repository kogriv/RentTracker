"""
Unit tests for ExcelGarageParser
"""

import pytest
from pathlib import Path
from decimal import Decimal
from datetime import date, datetime
from unittest.mock import Mock, patch, MagicMock

from src.parsers.file_parsers.excel_parser import ExcelGarageParser
from src.core.exceptions import ParseError


class TestExcelGarageParser:
    """Test cases for ExcelGarageParser"""
    
    @pytest.fixture
    def excel_parser(self):
        """Create ExcelGarageParser instance for testing"""
        return ExcelGarageParser()
    
    def test_get_supported_formats(self, excel_parser):
        """Test supported file formats"""
        formats = excel_parser.get_supported_formats()
        assert '.xlsx' in formats
        assert '.xls' in formats
    
    @patch('src.parsers.file_parsers.excel_parser.load_workbook')
    def test_validate_source_valid_file(self, mock_load_workbook, excel_parser):
        """Test source validation for valid Excel file"""
        mock_workbook = Mock()
        mock_load_workbook.return_value = mock_workbook
        
        test_file = Path("test.xlsx")
        
        with patch.object(test_file, 'exists', return_value=True):
            with patch.object(test_file, 'suffix', '.xlsx'):
                result = excel_parser.validate_source(test_file)
                assert result is True
        
        mock_workbook.close.assert_called_once()
    
    def test_validate_source_nonexistent_file(self, excel_parser):
        """Test source validation for nonexistent file"""
        test_file = Path("nonexistent.xlsx")
        result = excel_parser.validate_source(test_file)
        assert result is False
    
    def test_validate_source_wrong_extension(self, excel_parser):
        """Test source validation for wrong file extension"""
        test_file = Path("test.txt")
        
        with patch.object(test_file, 'exists', return_value=True):
            result = excel_parser.validate_source(test_file)
            assert result is False
    
    @patch('src.parsers.file_parsers.excel_parser.load_workbook')
    def test_validate_source_corrupted_file(self, mock_load_workbook, excel_parser):
        """Test source validation for corrupted Excel file"""
        mock_load_workbook.side_effect = Exception("Corrupted file")
        
        test_file = Path("corrupted.xlsx")
        
        with patch.object(test_file, 'exists', return_value=True):
            with patch.object(test_file, 'suffix', '.xlsx'):
                result = excel_parser.validate_source(test_file)
                assert result is False
    
    def test_is_header_row_valid_headers(self, excel_parser):
        """Test header row detection with valid headers"""
        # Mock row with Russian headers
        mock_row = [
            Mock(value="Гараж"),
            Mock(value="Сумма"),
            Mock(value="Первоначальная дата")
        ]
        
        result = excel_parser._is_header_row(mock_row)
        assert result is True
    
    def test_is_header_row_partial_headers(self, excel_parser):
        """Test header row detection with partial headers"""
        # Mock row with some headers
        mock_row = [
            Mock(value="Номер гаража"),
            Mock(value="Арендная плата"),
            Mock(value="Другое")
        ]
        
        result = excel_parser._is_header_row(mock_row)
        assert result is True  # Should match "гараж" and "сумма" keywords
    
    def test_is_header_row_no_headers(self, excel_parser):
        """Test header row detection with no headers"""
        mock_row = [
            Mock(value="1"),
            Mock(value="3500.00"),
            Mock(value="01.01.2025")
        ]
        
        result = excel_parser._is_header_row(mock_row)
        assert result is False
    
    def test_is_header_row_insufficient_columns(self, excel_parser):
        """Test header row detection with insufficient columns"""
        mock_row = [Mock(value="Гараж"), Mock(value="Сумма")]  # Only 2 columns
        
        result = excel_parser._is_header_row(mock_row)
        assert result is False
    
    def test_parse_garage_row_valid_data(self, excel_parser):
        """Test parsing valid garage row"""
        mock_row = [
            Mock(value="1"),
            Mock(value=3500.0),
            Mock(value=datetime(2025, 1, 15))
        ]
        
        garage_data = excel_parser._parse_garage_row(mock_row, 2)
        
        assert garage_data is not None
        assert garage_data['id'] == "1"
        assert garage_data['monthly_rent'] == Decimal("3500.0")
        assert garage_data['start_date'] == date(2025, 1, 15)
        assert garage_data['payment_day'] == 15
    
    def test_parse_garage_row_string_amount(self, excel_parser):
        """Test parsing garage row with string amount"""
        mock_row = [
            Mock(value="garage_2"),
            Mock(value="2,800.50"),  # String with comma
            Mock(value=datetime(2025, 2, 10))
        ]
        
        garage_data = excel_parser._parse_garage_row(mock_row, 3)
        
        assert garage_data is not None
        assert garage_data['id'] == "garage_2"
        assert garage_data['monthly_rent'] == Decimal("2800.50")
        assert garage_data['payment_day'] == 10
    
    def test_parse_garage_row_string_date(self, excel_parser):
        """Test parsing garage row with string date"""
        mock_row = [
            Mock(value="3"),
            Mock(value=4200.0),
            Mock(value="15.03.2025")  # String date
        ]
        
        garage_data = excel_parser._parse_garage_row(mock_row, 4)
        
        assert garage_data is not None
        assert garage_data['id'] == "3"
        assert garage_data['start_date'] == date(2025, 3, 15)
        assert garage_data['payment_day'] == 15
    
    def test_parse_garage_row_missing_data(self, excel_parser):
        """Test parsing garage row with missing data"""
        mock_row = [
            Mock(value="1"),
            Mock(value=None),  # Missing amount
            Mock(value=datetime(2025, 1, 15))
        ]
        
        garage_data = excel_parser._parse_garage_row(mock_row, 2)
        assert garage_data is None
    
    def test_parse_garage_row_invalid_amount(self, excel_parser):
        """Test parsing garage row with invalid amount"""
        mock_row = [
            Mock(value="1"),
            Mock(value="invalid_amount"),
            Mock(value=datetime(2025, 1, 15))
        ]
        
        garage_data = excel_parser._parse_garage_row(mock_row, 2)
        assert garage_data is None
    
    def test_parse_garage_row_negative_amount(self, excel_parser):
        """Test parsing garage row with negative amount"""
        mock_row = [
            Mock(value="1"),
            Mock(value=-3500.0),
            Mock(value=datetime(2025, 1, 15))
        ]
        
        garage_data = excel_parser._parse_garage_row(mock_row, 2)
        assert garage_data is None
    
    def test_parse_garage_row_invalid_date(self, excel_parser):
        """Test parsing garage row with invalid date"""
        mock_row = [
            Mock(value="1"),
            Mock(value=3500.0),
            Mock(value="invalid_date")
        ]
        
        garage_data = excel_parser._parse_garage_row(mock_row, 2)
        assert garage_data is None
    
    def test_parse_garage_row_insufficient_columns(self, excel_parser):
        """Test parsing garage row with insufficient columns"""
        mock_row = [Mock(value="1"), Mock(value=3500.0)]  # Only 2 columns
        
        garage_data = excel_parser._parse_garage_row(mock_row, 2)
        assert garage_data is None
    
    @patch('src.parsers.file_parsers.excel_parser.load_workbook')
    def test_parse_garages_successful(self, mock_load_workbook, excel_parser):
        """Test successful garage parsing"""
        # Mock workbook structure
        mock_workbook = Mock()
        mock_worksheet = Mock()
        mock_workbook.active = mock_worksheet
        mock_load_workbook.return_value = mock_workbook
        
        # Mock rows with header and data
        mock_rows = [
            # Header row
            [Mock(value="Гараж"), Mock(value="Сумма"), Mock(value="Дата")],
            # Data rows
            [Mock(value="1"), Mock(value=3500.0), Mock(value=datetime(2025, 1, 15))],
            [Mock(value="2"), Mock(value=2800.0), Mock(value=datetime(2025, 1, 16))],
        ]
        mock_worksheet.iter_rows.return_value = mock_rows
        
        test_file = Path("test.xlsx")
        
        with patch.object(excel_parser, 'validate_source', return_value=True):
            garages = excel_parser.parse_garages(test_file)
        
        assert len(garages) == 2
        assert garages[0]['id'] == "1"
        assert garages[1]['id'] == "2"
        mock_workbook.close.assert_called_once()
    
    @patch('src.parsers.file_parsers.excel_parser.load_workbook')
    def test_parse_garages_empty_rows(self, mock_load_workbook, excel_parser):
        """Test garage parsing with empty rows"""
        mock_workbook = Mock()
        mock_worksheet = Mock()
        mock_workbook.active = mock_worksheet
        mock_load_workbook.return_value = mock_workbook
        
        # Mock rows with empty rows
        mock_rows = [
            # Empty row
            [Mock(value=None), Mock(value=None), Mock(value=None)],
            # Header row
            [Mock(value="Гараж"), Mock(value="Сумма"), Mock(value="Дата")],
            # Empty row
            [Mock(value=None), Mock(value=None), Mock(value=None)],
            # Data row
            [Mock(value="1"), Mock(value=3500.0), Mock(value=datetime(2025, 1, 15))],
        ]
        mock_worksheet.iter_rows.return_value = mock_rows
        
        test_file = Path("test.xlsx")
        
        with patch.object(excel_parser, 'validate_source', return_value=True):
            garages = excel_parser.parse_garages(test_file)
        
        assert len(garages) == 1
        assert garages[0]['id'] == "1"
    
    def test_parse_garages_invalid_file(self, excel_parser):
        """Test garage parsing with invalid file"""
        test_file = Path("invalid.xlsx")
        
        with patch.object(excel_parser, 'validate_source', return_value=False):
            with pytest.raises(ParseError, match="Invalid Excel file"):
                excel_parser.parse_garages(test_file)
    
    @patch('src.parsers.file_parsers.excel_parser.load_workbook')
    def test_parse_garages_no_data(self, mock_load_workbook, excel_parser):
        """Test garage parsing with no valid data"""
        mock_workbook = Mock()
        mock_worksheet = Mock()
        mock_workbook.active = mock_worksheet
        mock_load_workbook.return_value = mock_workbook
        
        # Mock rows with only header
        mock_rows = [
            [Mock(value="Гараж"), Mock(value="Сумма"), Mock(value="Дата")],
        ]
        mock_worksheet.iter_rows.return_value = mock_rows
        
        test_file = Path("test.xlsx")
        
        with patch.object(excel_parser, 'validate_source', return_value=True):
            with pytest.raises(ParseError, match="No valid garage data found"):
                excel_parser.parse_garages(test_file)
    
    @patch('src.parsers.file_parsers.excel_parser.load_workbook')
    def test_parse_garages_workbook_error(self, mock_load_workbook, excel_parser):
        """Test garage parsing with workbook loading error"""
        from openpyxl.utils.exceptions import InvalidFileException
        mock_load_workbook.side_effect = InvalidFileException("Invalid file")
        
        test_file = Path("test.xlsx")
        
        with patch.object(excel_parser, 'validate_source', return_value=True):
            with pytest.raises(ParseError, match="Invalid Excel file format"):
                excel_parser.parse_garages(test_file)
    
    def test_parse_garage_row_various_date_formats(self, excel_parser):
        """Test parsing garage row with various date formats"""
        date_test_cases = [
            ("15.01.2025", date(2025, 1, 15)),
            ("15/01/2025", date(2025, 1, 15)),
            ("2025-01-15", date(2025, 1, 15)),
            ("15-01-2025", date(2025, 1, 15)),
        ]
        
        for date_str, expected_date in date_test_cases:
            mock_row = [
                Mock(value="1"),
                Mock(value=3500.0),
                Mock(value=date_str)
            ]
            
            garage_data = excel_parser._parse_garage_row(mock_row, 1)
            
            assert garage_data is not None, f"Failed to parse date: {date_str}"
            assert garage_data['start_date'] == expected_date
            assert garage_data['payment_day'] == expected_date.day
    
    def test_parse_garage_row_edge_case_amounts(self, excel_parser):
        """Test parsing garage row with edge case amounts"""
        amount_test_cases = [
            (0.01, Decimal("0.01")),
            (999999.99, Decimal("999999.99")),
            ("1,234.56", Decimal("1234.56")),
            ("5000,00", Decimal("5000.00")),  # European format
        ]
        
        for amount_input, expected_amount in amount_test_cases:
            mock_row = [
                Mock(value="1"),
                Mock(value=amount_input),
                Mock(value=datetime(2025, 1, 15))
            ]
            
            garage_data = excel_parser._parse_garage_row(mock_row, 1)
            
            assert garage_data is not None, f"Failed to parse amount: {amount_input}"
            assert garage_data['monthly_rent'] == expected_amount
